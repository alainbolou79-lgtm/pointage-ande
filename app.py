from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, make_response, send_from_directory
from functools import wraps
from datetime import datetime
import os, uuid, sqlite3, qrcode, io, threading, schedule, time
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pointage-secret-2024")

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pointage.db")
STRUCTURE_NOM = os.environ.get("STRUCTURE_NOM", "VOTRE STRUCTURE")
STRUCTURE_SERVICE = os.environ.get("STRUCTURE_SERVICE", "Service des Ressources Humaines")
CHEF_SERVICE_NOM = os.environ.get("CHEF_SERVICE_NOM", "Chef de Service")

# ── BASE DE DONNÉES ───────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    with get_conn() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS utilisateurs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL, prenom TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL, password TEXT NOT NULL,
                role TEXT DEFAULT 'admin', actif INTEGER DEFAULT 1,
                date_creation TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS agents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL, prenom TEXT NOT NULL,
                matricule TEXT UNIQUE NOT NULL,
                poste TEXT DEFAULT '', telephone TEXT DEFAULT '',
                actif INTEGER DEFAULT 1,
                date_creation TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS pointages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                agent_id INTEGER NOT NULL,
                date_pointage TEXT NOT NULL,
                heure_arrivee TEXT, heure_depart TEXT,
                duree_minutes INTEGER,
                FOREIGN KEY (agent_id) REFERENCES agents(id)
            );
            CREATE TABLE IF NOT EXISTS demandes_absence (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                agent_id INTEGER NOT NULL,
                date_debut TEXT NOT NULL, date_fin TEXT NOT NULL,
                nb_jours INTEGER NOT NULL, motif TEXT NOT NULL,
                type_absence TEXT DEFAULT 'conge',
                statut TEXT DEFAULT 'en_attente',
                commentaire_chef TEXT DEFAULT '',
                date_demande TEXT DEFAULT (datetime('now','localtime')),
                date_decision TEXT,
                FOREIGN KEY (agent_id) REFERENCES agents(id)
            );
            CREATE TABLE IF NOT EXISTS ip_autorisees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip_address TEXT UNIQUE NOT NULL,
                description TEXT DEFAULT 'Bureau',
                actif INTEGER DEFAULT 1,
                date_ajout TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS config (
                cle TEXT PRIMARY KEY, valeur TEXT
            );
        """)
        # Admin par défaut
        count = conn.execute("SELECT COUNT(*) FROM utilisateurs").fetchone()[0]
        if count == 0:
            conn.execute("""INSERT INTO utilisateurs (nom,prenom,email,password,role)
                VALUES (?,?,?,?,?)""",
                ("ADMIN","Système","admin@structure.ci",
                 generate_password_hash("admin2024"),"admin"))
        # Config par défaut
        for k,v in [("structure_nom", STRUCTURE_NOM),
                    ("structure_service", STRUCTURE_SERVICE),
                    ("chef_service", CHEF_SERVICE_NOM)]:
            conn.execute("INSERT OR IGNORE INTO config (cle,valeur) VALUES (?,?)",(k,v))
        conn.commit()

def get_config():
    with get_conn() as conn:
        return {r["cle"]:r["valeur"] for r in conn.execute("SELECT * FROM config").fetchall()}

def set_config(cle, valeur):
    with get_conn() as conn:
        conn.execute("INSERT OR REPLACE INTO config (cle,valeur) VALUES (?,?)",(cle,valeur))
        conn.commit()

# ── AUTH ──────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("user_role") != "admin":
            flash("Accès réservé aux administrateurs.","error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

@app.context_processor
def inject_config():
    cfg = get_config()
    return {"cfg": cfg, "current_user": session.get("user_nom",""),
            "current_role": session.get("user_role",""),
            "is_admin": session.get("user_role")=="admin"}

@app.route("/login", methods=["GET","POST"])
def login():
    if "user_id" in session: return redirect(url_for("dashboard"))
    if request.method == "POST":
        email = request.form.get("email","").strip()
        pwd   = request.form.get("password","").strip()
        with get_conn() as conn:
            u = conn.execute("SELECT * FROM utilisateurs WHERE email=? AND actif=1",(email,)).fetchone()
        if u and check_password_hash(u["password"], pwd):
            session["user_id"]   = u["id"]
            session["user_nom"]  = f"{u['prenom']} {u['nom']}"
            session["user_role"] = u["role"]
            return redirect(url_for("dashboard"))
        flash("Email ou mot de passe incorrect.","error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    date_sel = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    with get_conn() as conn:
        total_agents = conn.execute("SELECT COUNT(*) FROM agents WHERE actif=1").fetchone()[0]
        presents     = conn.execute("""SELECT COUNT(*) FROM pointages
            WHERE date_pointage=? AND heure_arrivee IS NOT NULL""",(date_sel,)).fetchone()[0]
        partis       = conn.execute("""SELECT COUNT(*) FROM pointages
            WHERE date_pointage=? AND heure_depart IS NOT NULL""",(date_sel,)).fetchone()[0]
        pointages    = [dict(r) for r in conn.execute("""
            SELECT p.*, a.nom, a.prenom, a.matricule, a.poste
            FROM pointages p JOIN agents a ON p.agent_id=a.id
            WHERE p.date_pointage=? ORDER BY p.heure_arrivee""",(date_sel,)).fetchall()]
        agents_ids   = {p["agent_id"] for p in pointages}
        tous_agents  = [dict(r) for r in conn.execute("SELECT * FROM agents WHERE actif=1").fetchall()]
        absents      = [a for a in tous_agents if a["id"] not in agents_ids]
        absences_att = conn.execute("""SELECT COUNT(*) FROM demandes_absence
            WHERE statut='en_attente'""").fetchone()[0]
    stats = {"total_agents":total_agents,"presents":presents,
             "absents":total_agents-presents,"partis":partis}
    return render_template("dashboard.html",
        stats=stats, pointages=pointages, absents=absents,
        date_sel=date_sel, absences_attente=absences_att)

# ── QR CODE ───────────────────────────────────────────────────────────────────
@app.route("/pointage/qr")
@login_required
def pointage_qr():
    token     = uuid.uuid4().hex
    pointages = []
    with get_conn() as conn:
        today = datetime.now().strftime("%Y-%m-%d")
        rows  = conn.execute("""SELECT p.*, a.nom, a.prenom, a.matricule
            FROM pointages p JOIN agents a ON p.agent_id=a.id
            WHERE p.date_pointage=? ORDER BY p.heure_arrivee DESC""",(today,)).fetchall()
        pointages = [dict(r) for r in rows]
    base_url = request.host_url.rstrip('/')
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    return render_template("pointage_qr.html",
        token=token, pointages=pointages, base_url=base_url, now=now)

@app.route("/pointage/qr-image")
def pointage_qr_image():
    token    = request.args.get("token","")
    base_url = request.host_url.rstrip('/')
    url      = f"{base_url}/mobile?token={token}"
    qr = qrcode.QRCode(version=1, box_size=8, border=3)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Type"]  = "image/png"
    response.headers["Cache-Control"] = "no-cache"
    return response

@app.route("/pointage/new-token")
def new_token():
    return jsonify({"token": uuid.uuid4().hex})

@app.route("/pointage/recent")
def pointage_recent():
    today = datetime.now().strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute("""SELECT p.*, a.nom, a.prenom, a.matricule
            FROM pointages p JOIN agents a ON p.agent_id=a.id
            WHERE p.date_pointage=? ORDER BY p.heure_arrivee DESC""",(today,)).fetchall()
    return jsonify([dict(r) for r in rows])

# ── MOBILE ────────────────────────────────────────────────────────────────────
@app.route("/mobile")
def mobile():
    return render_template("mobile.html")

@app.route("/installer")
def installer():
    return render_template("installer.html")

@app.route("/api/agent/<matricule>")
def get_agent(matricule):
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM agents WHERE matricule=? AND actif=1",
                        (matricule.upper(),)).fetchone()
    if r:
        a = dict(r)
        role = "chef_de_service" if "chef" in a.get("poste","").lower() or \
               a.get("poste","").startswith("CF_") else "agent"
        return jsonify({"found":True,"nom":a["nom"],"prenom":a["prenom"],
                        "poste":a["poste"],"role":role})
    return jsonify({"found":False})

@app.route("/api/pointer", methods=["POST"])
def pointer():
    data      = request.get_json()
    matricule = data.get("matricule","").upper()
    type_p    = data.get("type","arrivee")

    # Vérif IP
    ip = request.headers.get('X-Forwarded-For',
         request.headers.get('X-Real-IP', request.remote_addr))
    if ip and ',' in ip: ip = ip.split(',')[0].strip()

    with get_conn() as conn:
        ips_ok = [dict(r) for r in conn.execute(
            "SELECT * FROM ip_autorisees WHERE actif=1").fetchall()]
        if ips_ok and not any(i["ip_address"]==ip for i in ips_ok):
            return jsonify({"success":False,"icon":"🚫",
                "message":"Pointage non autorisé !",
                "detail":f"Vous devez être connecté au WiFi du bureau. (IP: {ip})",
                "heure":""})

        agent = conn.execute("SELECT * FROM agents WHERE matricule=? AND actif=1",
                            (matricule,)).fetchone()
        if not agent:
            return jsonify({"success":False,"icon":"❌",
                "message":"Matricule introuvable","detail":"","heure":""})

        agent = dict(agent)
        today = datetime.now().strftime("%Y-%m-%d")
        heure = datetime.now().strftime("%H:%M:%S")

        if type_p == "arrivee":
            existing = conn.execute("""SELECT * FROM pointages
                WHERE agent_id=? AND date_pointage=?""",(agent["id"],today)).fetchone()
            if existing and existing["heure_arrivee"]:
                return jsonify({"success":False,"warning":True,"icon":"⚠️",
                    "message":"Déjà pointé ce matin",
                    "heure":existing["heure_arrivee"][:5],"detail":""})
            if existing:
                conn.execute("UPDATE pointages SET heure_arrivee=? WHERE id=?",
                            (heure, existing["id"]))
            else:
                conn.execute("""INSERT INTO pointages
                    (agent_id,date_pointage,heure_arrivee) VALUES (?,?,?)""",
                    (agent["id"],today,heure))
            conn.commit()
            return jsonify({"success":True,"icon":"✅",
                "message":f"Bonjour {agent['prenom']} !",
                "heure":heure[:5],"detail":"Arrivée enregistrée ✅"})
        else:
            existing = conn.execute("""SELECT * FROM pointages
                WHERE agent_id=? AND date_pointage=?""",(agent["id"],today)).fetchone()
            if not existing or not existing["heure_arrivee"]:
                return jsonify({"success":False,"icon":"❌",
                    "message":"Arrivée non enregistrée","heure":"","detail":""})
            if existing["heure_depart"]:
                return jsonify({"success":False,"warning":True,"icon":"⚠️",
                    "message":"Départ déjà enregistré",
                    "heure":existing["heure_depart"][:5],"detail":""})
            arr = datetime.strptime(existing["heure_arrivee"],"%H:%M:%S")
            dep = datetime.strptime(heure,"%H:%M:%S")
            duree = int((dep-arr).total_seconds()/60)
            conn.execute("""UPDATE pointages SET heure_depart=?, duree_minutes=?
                WHERE id=?""",(heure, duree, existing["id"]))
            conn.commit()
            return jsonify({"success":True,"icon":"👋",
                "message":f"Au revoir {agent['prenom']} !",
                "heure":heure[:5],"detail":"Départ enregistré ✅"})

# ── ABSENCES ──────────────────────────────────────────────────────────────────
@app.route("/absences")
@login_required
def absences():
    with get_conn() as conn:
        attente  = [dict(r) for r in conn.execute("""
            SELECT d.*, a.nom, a.prenom, a.matricule, a.poste
            FROM demandes_absence d JOIN agents a ON d.agent_id=a.id
            WHERE d.statut='en_attente' ORDER BY d.date_demande DESC""").fetchall()]
        historique = [dict(r) for r in conn.execute("""
            SELECT d.*, a.nom, a.prenom, a.matricule
            FROM demandes_absence d JOIN agents a ON d.agent_id=a.id
            WHERE d.statut!='en_attente' ORDER BY d.date_decision DESC LIMIT 50""").fetchall()]
    return render_template("absences.html", attente=attente, historique=historique)

@app.route("/absences/<int:did>/decider", methods=["POST"])
@login_required
def decider_absence(did):
    decision    = request.form.get("decision","")
    commentaire = request.form.get("commentaire","")
    now         = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute("""UPDATE demandes_absence SET statut=?,
            commentaire_chef=?, date_decision=? WHERE id=?""",
            (decision, commentaire, now, did))
        conn.commit()
    flash(f"✅ Demande {'accordée' if decision=='accorde' else 'refusée'} !","success")
    return redirect(url_for("absences"))

@app.route("/api/absences/soumettre", methods=["POST"])
def soumettre_absence():
    data      = request.get_json()
    matricule = data.get("matricule","").upper()
    with get_conn() as conn:
        agent = conn.execute("SELECT * FROM agents WHERE matricule=? AND actif=1",
                            (matricule,)).fetchone()
        if not agent:
            return jsonify({"success":False,"message":"Matricule introuvable"})
        conn.execute("""INSERT INTO demandes_absence
            (agent_id,date_debut,date_fin,nb_jours,motif,type_absence)
            VALUES (?,?,?,?,?,?)""",
            (agent["id"],data.get("date_debut"),data.get("date_fin"),
             data.get("nb_jours",1),data.get("motif",""),
             data.get("type_absence","conge")))
        conn.commit()
    return jsonify({"success":True,"message":"Demande soumise !"})

@app.route("/api/absences/mes-demandes")
def mes_absences():
    matricule = request.args.get("matricule","").upper()
    with get_conn() as conn:
        agent = conn.execute("SELECT * FROM agents WHERE matricule=?",
                            (matricule,)).fetchone()
        if not agent:
            return jsonify([])
        rows = conn.execute("""SELECT * FROM demandes_absence
            WHERE agent_id=? ORDER BY date_demande DESC LIMIT 20""",
            (agent["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/absences/chef")
def absences_chef():
    with get_conn() as conn:
        attente = [dict(r) for r in conn.execute("""
            SELECT d.*, a.nom, a.prenom, a.matricule, a.poste
            FROM demandes_absence d JOIN agents a ON d.agent_id=a.id
            WHERE d.statut='en_attente' ORDER BY d.date_demande DESC""").fetchall()]
        historique = [dict(r) for r in conn.execute("""
            SELECT d.*, a.nom, a.prenom, a.matricule
            FROM demandes_absence d JOIN agents a ON d.agent_id=a.id
            WHERE d.statut!='en_attente' ORDER BY d.date_decision DESC LIMIT 20""").fetchall()]
        stats = {
            "en_attente": conn.execute("SELECT COUNT(*) FROM demandes_absence WHERE statut='en_attente'").fetchone()[0],
            "accordees":  conn.execute("SELECT COUNT(*) FROM demandes_absence WHERE statut='accorde'").fetchone()[0],
        }
    return jsonify({"en_attente":attente,"historique":historique,"stats":stats})

@app.route("/api/absences/decider", methods=["POST"])
def decider_absence_mobile():
    data        = request.get_json()
    did         = data.get("did")
    decision    = data.get("decision","")
    commentaire = data.get("commentaire","")
    now         = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute("""UPDATE demandes_absence SET statut=?,
            commentaire_chef=?, date_decision=? WHERE id=?""",
            (decision, commentaire, now, did))
        conn.commit()
    return jsonify({"success":True})

# ── AGENTS ────────────────────────────────────────────────────────────────────
@app.route("/agents")
@login_required
def agents():
    with get_conn() as conn:
        liste = [dict(r) for r in conn.execute(
            "SELECT * FROM agents WHERE actif=1 ORDER BY nom").fetchall()]
    return render_template("agents.html", agents=liste)

@app.route("/agents/ajouter", methods=["POST"])
@admin_required
def ajouter_agent():
    with get_conn() as conn:
        try:
            conn.execute("""INSERT INTO agents (nom,prenom,matricule,poste,telephone,actif)
                VALUES (?,?,?,?,?,1)""",
                (request.form.get("nom","").strip(),
                 request.form.get("prenom","").strip(),
                 request.form.get("matricule","").strip().upper(),
                 request.form.get("poste","").strip(),
                 request.form.get("telephone","").strip()))
            conn.commit()
            flash("✅ Agent ajouté !","success")
        except Exception as e:
            flash(f"❌ Erreur : {e}","error")
    return redirect(url_for("agents"))

@app.route("/agents/<int:aid>/supprimer", methods=["POST"])
@admin_required
def supprimer_agent(aid):
    with get_conn() as conn:
        conn.execute("UPDATE agents SET actif=0 WHERE id=?",(aid,))
        conn.commit()
    flash("Agent supprimé.","success")
    return redirect(url_for("agents"))

# ── IP CONFIG ─────────────────────────────────────────────────────────────────
@app.route("/config-ip")
@admin_required
def config_ip():
    ip_actuelle = request.headers.get('X-Forwarded-For',
                  request.headers.get('X-Real-IP', request.remote_addr))
    if ip_actuelle and ',' in ip_actuelle:
        ip_actuelle = ip_actuelle.split(',')[0].strip()
    with get_conn() as conn:
        ips = [dict(r) for r in conn.execute("SELECT * FROM ip_autorisees WHERE actif=1").fetchall()]
    return render_template("config_ip.html", ips=ips, ip_actuelle=ip_actuelle)

@app.route("/config-ip/ajouter-auto", methods=["POST"])
@admin_required
def ajouter_ip_auto():
    ip = request.headers.get('X-Forwarded-For',
         request.headers.get('X-Real-IP', request.remote_addr))
    if ip and ',' in ip: ip = ip.split(',')[0].strip()
    with get_conn() as conn:
        try:
            conn.execute("INSERT INTO ip_autorisees (ip_address,description) VALUES (?,?)",
                        (ip, request.form.get("description","Bureau")))
            conn.commit()
            flash(f"✅ IP {ip} enregistrée !","success")
        except:
            flash(f"⚠️ IP {ip} déjà enregistrée.","warning")
    return redirect(url_for("config_ip"))

@app.route("/config-ip/supprimer/<int:iid>", methods=["POST"])
@admin_required
def supprimer_ip(iid):
    with get_conn() as conn:
        conn.execute("DELETE FROM ip_autorisees WHERE id=?",(iid,))
        conn.commit()
    flash("IP supprimée.","success")
    return redirect(url_for("config_ip"))

@app.route("/mon-ip")
def mon_ip():
    ip = request.headers.get('X-Forwarded-For',
         request.headers.get('X-Real-IP', request.remote_addr))
    if ip and ',' in ip: ip = ip.split(',')[0].strip()
    with get_conn() as conn:
        ips_ok = [dict(r) for r in conn.execute("SELECT * FROM ip_autorisees WHERE actif=1").fetchall()]
    autorisee = not ips_ok or any(i["ip_address"]==ip for i in ips_ok)
    return jsonify({"ip":ip,"autorisee":autorisee})

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
@app.route("/configuration", methods=["GET","POST"])
@admin_required
def configuration():
    if request.method == "POST":
        for k in ["structure_nom","structure_service","chef_service"]:
            v = request.form.get(k,"").strip()
            if v: set_config(k,v)
        flash("✅ Configuration sauvegardée !","success")
        return redirect(url_for("configuration"))
    return render_template("configuration.html", cfg=get_config())

# ── EXPORT ────────────────────────────────────────────────────────────────────
@app.route("/export/pointage")
@login_required
def export_pointage():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    date_sel = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    with get_conn() as conn:
        pointages = [dict(r) for r in conn.execute("""
            SELECT p.*, a.nom, a.prenom, a.matricule, a.poste
            FROM pointages p JOIN agents a ON p.agent_id=a.id
            WHERE p.date_pointage=? ORDER BY p.heure_arrivee""",(date_sel,)).fetchall()]
        stats = {
            "total":    conn.execute("SELECT COUNT(*) FROM agents WHERE actif=1").fetchone()[0],
            "presents": conn.execute("SELECT COUNT(*) FROM pointages WHERE date_pointage=? AND heure_arrivee IS NOT NULL",(date_sel,)).fetchone()[0],
        }
    cfg     = get_config()
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Pointage"
    ws.sheet_view.showGridLines = False
    thin   = Side(style="thin",color="CCCCCC")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    # Titre
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{cfg.get('structure_nom','STRUCTURE')} – FEUILLE DE PRÉSENCE DU {date_sel}"
    ws["A1"].font = Font(bold=True,size=13,color="FFFFFF")
    ws["A1"].fill = PatternFill("solid",fgColor="1A6B2A")
    ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Présents: {stats['presents']} | Générée le {now_str}"
    ws["A2"].font = Font(italic=True,size=10,color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    headers = ["MATRICULE","NOM COMPLET","POSTE","ARRIVÉE","DÉPART","DURÉE","STATUT"]
    for col,h in enumerate(headers,1):
        cell = ws.cell(row=4,column=col,value=h)
        cell.font = Font(bold=True,color="FFFFFF",size=10)
        cell.fill = PatternFill("solid",fgColor="1A6B2A")
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.border = border
    ws.row_dimensions[4].height = 22
    for i,p in enumerate(pointages):
        row = 5+i
        bg  = "F0FFF0" if i%2==0 else "FFFFFF"
        duree_str = ""
        if p.get("duree_minutes"):
            h = p["duree_minutes"]//60; m = p["duree_minutes"]%60
            duree_str = f"{h}h{m:02d}"
        statut = "Parti" if p.get("heure_depart") else "Présent" if p.get("heure_arrivee") else "Absent"
        for col,val in enumerate([
            p["matricule"],f"{p['prenom']} {p['nom']}",p.get("poste",""),
            p.get("heure_arrivee","")[:5] if p.get("heure_arrivee") else "—",
            p.get("heure_depart","")[:5] if p.get("heure_depart") else "—",
            duree_str or "—",statut
        ],1):
            cell = ws.cell(row=row,column=col,value=val)
            cell.fill = PatternFill("solid",fgColor=bg)
            cell.border = border
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center",horizontal="center")
        ws.row_dimensions[row].height = 18
    for i,w in enumerate([14,24,20,12,12,10,12],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Signature
    sig_row = ws.max_row + 3
    ws.merge_cells(f"E{sig_row}:G{sig_row}")
    ws[f"E{sig_row}"] = cfg.get("chef_service","Chef de Service")
    ws[f"E{sig_row}"].font = Font(bold=True,size=11)
    ws[f"E{sig_row}"].alignment = Alignment(horizontal="center")
    sig_row2 = sig_row + 1
    ws.merge_cells(f"E{sig_row2}:G{sig_row2}")
    ws[f"E{sig_row2}"] = cfg.get("structure_service","")
    ws[f"E{sig_row2}"].alignment = Alignment(horizontal="center")
    output = io.BytesIO(); wb.save(output); output.seek(0)
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=Pointage_{date_sel}.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

# ── MAIN ──────────────────────────────────────────────────────────────────────
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── IMPORT AGENTS DEPUIS EXCEL ────────────────────────────────────────────────
@app.route("/agents/importer", methods=["GET","POST"])
@admin_required
def importer_agents():
    if request.method == "POST":
        import openpyxl, io as io_buf
        f = request.files.get("fichier")
        if not f:
            flash("❌ Aucun fichier sélectionné.","error")
            return redirect(url_for("importer_agents"))
        try:
            wb = openpyxl.load_workbook(io_buf.BytesIO(f.read()), read_only=True)
            ws = wb.active
            count = 0
            with get_conn() as conn:
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    if not row[0]: continue
                    nom       = str(row[0]).strip() if row[0] else ''
                    prenom    = str(row[1]).strip() if len(row)>1 and row[1] else ''
                    matricule = str(row[2]).strip().upper() if len(row)>2 and row[2] else ''
                    poste     = str(row[3]).strip() if len(row)>3 and row[3] else ''
                    telephone = str(row[4]).strip() if len(row)>4 and row[4] else ''
                    if not matricule: continue
                    try:
                        conn.execute("""INSERT OR REPLACE INTO agents 
                            (nom,prenom,matricule,poste,telephone,actif)
                            VALUES (?,?,?,?,?,1)""",
                            (nom,prenom,matricule,poste,telephone))
                        count += 1
                    except: pass
                conn.commit()
            flash(f"✅ {count} agent(s) importé(s) avec succès !","success")
        except Exception as e:
            flash(f"❌ Erreur : {e}","error")
        return redirect(url_for("agents"))
    return render_template("importer_agents.html")

# ── CONTROLE APPAREILS ────────────────────────────────────────────────────────
@app.route("/appareils")
@admin_required
def appareils():
    today = datetime.now().strftime("%Y-%m-%d")
    with get_conn() as conn:
        # Créer les tables si elles n'existent pas
        conn.execute("""CREATE TABLE IF NOT EXISTS appareils_pointes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_pointage TEXT NOT NULL,
            mac_address TEXT DEFAULT '',
            ip_address TEXT DEFAULT '',
            agent_id INTEGER,
            type_pointage TEXT DEFAULT 'arrivee',
            date_creation TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS appareils_exclus (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mac_address TEXT DEFAULT '',
            ip_address TEXT DEFAULT '',
            description TEXT DEFAULT '',
            actif INTEGER DEFAULT 1,
            date_ajout TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
        try:
            liste = [dict(r) for r in conn.execute("""
                SELECT a.*, ag.nom, ag.prenom, ag.matricule
                FROM appareils_pointes a
                JOIN agents ag ON a.agent_id = ag.id
                WHERE a.date_pointage = ?
                ORDER BY a.date_creation DESC
            """, (today,)).fetchall()]
        except: liste = []
    return render_template("appareils.html", appareils=liste, today=today)

# ── EXCLUSIONS MAC ────────────────────────────────────────────────────────────
@app.route("/appareils/exclure", methods=["POST"])
@admin_required
def exclure_appareil():
    mac = request.form.get("mac_address","").strip()
    ip  = request.form.get("ip_address","").strip()
    desc= request.form.get("description","Appareil test").strip()
    with get_conn() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS appareils_exclus (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mac_address TEXT DEFAULT '',
            ip_address  TEXT DEFAULT '',
            description TEXT DEFAULT '',
            actif       INTEGER DEFAULT 1,
            date_ajout  TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.execute("""INSERT INTO appareils_exclus 
            (mac_address, ip_address, description) VALUES (?,?,?)""",
            (mac, ip, desc))
        conn.commit()
    flash(f"✅ Appareil exclu : {mac or ip}","success")
    return redirect(url_for("appareils"))

@app.route("/appareils/exclus")
@admin_required
def appareils_exclus():
    with get_conn() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS appareils_exclus (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mac_address TEXT DEFAULT '',
            ip_address  TEXT DEFAULT '',
            description TEXT DEFAULT '',
            actif       INTEGER DEFAULT 1,
            date_ajout  TEXT DEFAULT (datetime('now','localtime'))
        )""")
        liste = [dict(r) for r in conn.execute(
            "SELECT * FROM appareils_exclus ORDER BY date_ajout DESC").fetchall()]
    return render_template("appareils_exclus.html", exclus=liste)

@app.route("/appareils/exclus/<int:eid>/toggle", methods=["POST"])
@admin_required
def toggle_exclusion(eid):
    with get_conn() as conn:
        r = conn.execute("SELECT actif FROM appareils_exclus WHERE id=?", (eid,)).fetchone()
        if r:
            conn.execute("UPDATE appareils_exclus SET actif=? WHERE id=?",
                        (0 if r["actif"] else 1, eid))
            conn.commit()
    flash("✅ Statut mis à jour !","success")
    return redirect(url_for("appareils_exclus"))

@app.route("/appareils/exclus/<int:eid>/supprimer", methods=["POST"])
@admin_required
def supprimer_exclusion(eid):
    with get_conn() as conn:
        conn.execute("DELETE FROM appareils_exclus WHERE id=?", (eid,))
        conn.commit()
    flash("Exclusion supprimée.","success")
    return redirect(url_for("appareils_exclus"))
